import numpy as np

from tifffile import TiffFile
import matplotlib.pyplot as plt

import re
from pathlib import Path
import openpyxl
import re
import random

import torch
torch.manual_seed(17)
import torchvision


from framework.dataset import LandCoverData as LCD    
    
def AinterB(masques_cal,mask):
    masques_inter_mask = np.zeros(10)
    pmask = np.zeros(10)
    pmasques_cal = np.zeros(10)
    for i in range(10):
        a = np.ndarray.flatten(masques_cal) == i
        b = np.ndarray.flatten(mask) == i
        masques_inter_mask[i] += np.sum([all(tup) for tup in zip(a, b)])
        pmasques_cal[i] += np.sum(a)
        pmask[i] += np.sum(b)
    masques_union_mask = pmask + pmasques_cal - masques_inter_mask
    proba = masques_inter_mask/masques_union_mask
    proba = np.nan_to_num(proba)
    return proba

def ecrire_excel(mask_cal, nom):
    """créé un fichier excel pour la soumission.
    Args:
       mask_cal (dictionnaire dont les clef sont le chemin des images): ex : dataset/test/images/  .tif 
       nom (string permettant de nommer le fichier excel)
    """
    nom_fichier = str(nom+'.xlsx')
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = "sample_id"
    sheet['B1'] = "no_data"
    sheet['C1'] = "clouds"
    sheet['D1'] = "artificial"
    sheet['E1'] = "cultivated"
    sheet['F1'] = "broadleaf"
    sheet['G1'] = "coniferous"
    sheet['H1'] = "herbaceous"
    sheet['I1'] = "natural"
    sheet['J1'] = "snow"
    sheet['K1'] = "water"
    for k, batch in enumerate(mask_cal):
        #print('batch', batch)
        idx = int(re.findall('\d+', batch)[0]) 
        mask = np.array(mask_cal[f"dataset/test/images/{idx}.tif"])
        no_data = np.count_nonzero(mask == 0)/(mask.shape[0]*mask.shape[1])
        clouds = np.count_nonzero(mask == 1)/(mask.shape[0]*mask.shape[1])
        artificial = np.count_nonzero(mask ==2)/(mask.shape[0]*mask.shape[1])
        cultivated = np.count_nonzero(mask == 3)/(mask.shape[0]*mask.shape[1])
        broadleaf = np.count_nonzero(mask == 4)/(mask.shape[0]*mask.shape[1])
        coniferous = np.count_nonzero(mask == 5)/(mask.shape[0]*mask.shape[1])
        herbaceous = np.count_nonzero(mask == 6)/(mask.shape[0]*mask.shape[1])
        natural = np.count_nonzero(mask == 7)/(mask.shape[0]*mask.shape[1])
        snow = np.count_nonzero(mask == 8)/(mask.shape[0]*mask.shape[1])
        water = np.count_nonzero(mask == 9)/(mask.shape[0]*mask.shape[1])
        sheet.cell(k+2,1).value = idx
        sheet.cell(k+2,2).value = str(no_data).replace(',', '.')
        sheet.cell(k+2,3).value = str(clouds).replace(',', '.')
        sheet.cell(k+2,4).value = str(artificial).replace(',', '.')
        sheet.cell(k+2,5).value = str(cultivated).replace(',', '.')
        sheet.cell(k+2,6).value = str(broadleaf).replace(',', '.')
        sheet.cell(k+2,7).value = str(coniferous).replace(',', '.')
        sheet.cell(k+2,8).value = str(herbaceous).replace(',', '.')
        sheet.cell(k+2,9).value = str(natural).replace(',', '.')
        sheet.cell( k+2,10).value = str(snow).replace(',', '.')
        sheet.cell( k+2,11).value = str(water).replace(',', '.')
    workbook.save(nom_fichier)
    workbook.close()
    print("done")


def show_all(image_path=False, mask_path=False, mask_cal=False, idx=False, legend_masque=False):
    ax=0
    axx=0
    
    if image_path and ((idx<18491 and idx>16811) or (idx<13449 and idx > 10086)):
        exemple = 'test'
        image_path = f'dataset/test/images/{idx}.tif'
        with TiffFile(image_path) as tif:
            img = tif.asarray()  
        ax += 1
       


    #id_img = int(re.findall('\d+', idx)[0])

    if ((idx<10087 and idx>1) or (idx<16810 and idx > 13449) or idx>18942):
        exemple = 'train'
        if image_path:
            image_path = f'dataset/train/images/{idx}.tif'
            with TiffFile(image_path) as tif:
                img = tif.asarray()
            ax += 1
        if mask_path:
            mask_path = f'dataset/train/masks/{idx}.tif'
            with TiffFile(mask_path) as tif:
                masque = tif.asarray()
            ax += 1

    if mask_cal:    
        ax += 1
        masque_cal = mask_cal[f"dataset/{exemple}/images/{idx}.tif"]
    fig, axs = plt.subplots(1, ax, figsize=(30, 30))
    classes_colorpalette = {c: color/255. for (c, color) in LCD.CLASSES_COLORPALETTE.items()}
    if img.any():
        show_image(img, display_min=0, display_max=2200, ax=axs[axx])
        axs[axx].set_title(f'Image: {idx}')

        axx+=1
    if mask_path:
        show_mask(masque, classes_colorpalette = classes_colorpalette, classes=LCD.CLASSES, add_legend=legend_masque, ax=axs[axx])
        axs[axx].set_title(f'Mask: {idx}')
        axx+=1
    if mask_cal:
        show_mask(masque_cal, classes_colorpalette = classes_colorpalette, classes=LCD.CLASSES, add_legend=legend_masque, ax=axs[axx])
        axs[axx].set_title(f'Ton Masque: {idx}')

def show_image(image, display_min=50, display_max=400, ax=None):
    """Show an image.
    Args:
        image (numpy.array[uint16]): the image. If the image is 16-bit, apply bytescaling to convert to 8-bit
    """
    if image.dtype == np.uint16:
        iscale = display_max - display_min
        scale = 255 / iscale
        byte_im = (image) * scale
        byte_im = (byte_im.clip(0, 255) + 0.5).astype(np.uint8)
        image = byte_im
    # show image
   
    if ax is None:
        fig, ax = plt.subplots(figsize=(10, 10))
    ax.axis("off")
    im = ax.imshow(image)
    return im


def show_mask(mask, classes_colorpalette, classes=None, add_legend=True, ax=None):
    """Show a a semantic segmentation mask.
    Args:
       mask (numpy.array[uint8]): the mask in 8-bit
       classes_colorpalette (dict[int, tuple]): dict mapping class index to an RGB color in [0, 1]
       classes (list[str], optional): list of class labels
       add_legend
    """
    show_mask = np.empty((*mask.shape, 3))
    for c, color in classes_colorpalette.items():
        show_mask[mask == c, :] = color
    if ax is None:
        fig, ax = plt.subplots(figsize=(10, 10))
    ax.axis("off")
    im = ax.imshow(show_mask)
    if add_legend:
        # show legend mapping pixel colors to class names
        import matplotlib.patches as mpatches
        handles = []
        for c, color in classes_colorpalette.items():
            handles.append(mpatches.Patch(color=color, label=classes[c]))
        ax.legend(handles=handles)
    return im

def compter_prob(masque):
    u = masque.flatten()
    y_true = np.zeros(10)
    a =np.bincount(u)
    count =a/np.sum(a)
    y_true[0:len(count)]= count
    return(y_true)

def compter_classes(masque):
    u = masque.flatten()
    y_true = np.zeros(10)
    a =np.bincount(u)
    count =a
    y_true[0:len(count)]= count
    return(y_true)

# petit test graphique
#iterator = iter(train_loader)
#data = next(iterator)

#arr = np.array(data['image'][2]).astype(np.uint16)
#fig, axs = plt.subplots(1, 1, figsize=(10, 10))
#show_image(arr, display_min=0, display_max=2200, ax=axs)
def Data_augmentation(mon_image, mon_masque):
    a = random.randint(0, 3)          
    image = torchvision.transforms.functional.rotate(mon_image, a*90)
    masques = torchvision.transforms.functional.rotate(mon_masque, a*90)

    return image, masques